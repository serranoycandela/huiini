#-*- encoding: utf-8 -*-
from PySide2.QtCore import *
from PySide2.QtCore import Qt
from PySide2.QtGui import *
from PySide2 import QtGui, QtCore, QtWidgets
from PySide2.QtWidgets import QTableWidgetItem, QFileDialog, QProgressDialog, QMessageBox, QListView, QAbstractItemView, QTreeView
import sys
import guiV2
from os import listdir
from os.path import isfile, join, basename
import shutil
import os
import win32print
import win32api
import time as time_old
from subprocess import Popen
from FacturasLocal import FacturaLocal as Factura
import math
import json
#import xlsxwriter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from openpyxl import load_workbook,  Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


from datetime import datetime





##C:\Python36\Scripts\pyside2-uic.exe mainwindowV2.ui -o guiV2.py
##C:\Python36\Scripts\pyinstaller.exe huiini.py





try:
    scriptDirectory = os.path.dirname(os.path.abspath(__file__))
except NameError:  # We are the main py2exe script, not a module
    scriptDirectory = os.path.dirname(os.path.abspath(sys.argv[0]))



class ImgWidgetPalomita(QtWidgets.QLabel):

    def __init__(self, parent=None):
        super(ImgWidgetPalomita, self).__init__(parent)
        pic_palomita = QtGui.QPixmap(join(scriptDirectory,"palomita.png"))
        self.setPixmap(pic_palomita)

class ImgWidgetTache(QtWidgets.QLabel):

    def __init__(self, parent=None):
        super(ImgWidgetTache, self).__init__(parent)
        pic_tache = QtGui.QPixmap(join(scriptDirectory,"x.png"))
        self.setPixmap(pic_tache)



class Ui_MainWindow(QtWidgets.QMainWindow, guiV2.Ui_MainWindow):

    def __init__(self, parent=None):
        super(Ui_MainWindow, self).__init__(parent)
        self.setupUi(self)

        print(scriptDirectory)
        logoPix = QtGui.QPixmap(join(scriptDirectory,"logo.png"))
        self.labelLogo.setPixmap(logoPix)
        self.pdflatex_path = "C:/Program Files/MiKTeX 2.9/miktex/bin/x64/pdflatex.exe"

        self.carpetaChooser.clicked.connect(self.cualCarpeta)
        #self.descarga_bt.clicked.connect(self.descarga_mesta)
        self.imprimir.clicked.connect(self.imprime)

        self.impresora.clicked.connect(self.cambiaImpresora)
        self.botonCancela.clicked.connect(self.cancelaImpresion)

        self.listaDeImpresoras.currentItemChanged.connect(self.cambiaSeleccionDeImpresora)

        self.tableWidget_xml.setColumnCount(16)
        self.tableWidget_xml.setColumnWidth(0,30)#pdf
        self.tableWidget_xml.setColumnWidth(1,95)#fecha
        self.tableWidget_xml.setColumnWidth(2,70)#uuid
        self.tableWidget_xml.setColumnWidth(3,120)#receptor-nombre
        self.tableWidget_xml.setColumnWidth(4,120)#emisor-rfc
        self.tableWidget_xml.setColumnWidth(5,120)#concepto
        self.tableWidget_xml.setColumnWidth(6,30)#version
        self.tableWidget_xml.setColumnWidth(7,75)#Subtotal
        self.tableWidget_xml.setColumnWidth(8,80)#Descuento
        self.tableWidget_xml.setColumnWidth(9,80)#traslados-iva
        self.tableWidget_xml.setColumnWidth(10,80)#traslados-ieps
        self.tableWidget_xml.setColumnWidth(11,75)#retIVA
        self.tableWidget_xml.setColumnWidth(12,75)#retISR
        self.tableWidget_xml.setColumnWidth(13,80)#total
        self.tableWidget_xml.setColumnWidth(14,74)#formaDePago
        self.tableWidget_xml.setColumnWidth(15,77)#metodoDePago

        self.tableWidget_xml.verticalHeader().setFixedWidth(35)

        self.tableWidget_resumen.setColumnCount(10)
        self.tableWidget_resumen.setColumnWidth(0,30)
        self.tableWidget_resumen.setColumnWidth(1,152)
        self.tableWidget_resumen.setColumnWidth(2,192)
        self.tableWidget_resumen.setColumnWidth(3,80)
        self.tableWidget_resumen.setColumnWidth(4,80)
        self.tableWidget_resumen.setColumnWidth(5,80)
        self.tableWidget_resumen.setColumnWidth(6,80)
        self.tableWidget_resumen.setColumnWidth(7,65)
        self.tableWidget_resumen.setColumnWidth(8,65)
        self.tableWidget_resumen.setColumnWidth(9,80)
        self.tableWidget_resumen.setRowCount(2)
        #self.tableWidget_resumen.verticalHeader().setFixedWidth(35)

        header = self.tableWidget_xml.verticalHeader()
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(self.handleHeaderMenu)

        self.ponEncabezado()

        self.tableWidget_xml.cellDoubleClicked.connect(self.meDoblePicaronXML)
        self.tableWidget_resumen.cellDoubleClicked.connect(self.meDoblePicaronResumen)

    def as_text(self,value):
        if value is None:
            return ""
        return str(value)

    def style_ws(self, ws, columna_totales, sumas_row):
        cell_border = Border(left=Side(border_style='medium', color='FF000000'),
                     right=Side(border_style='medium', color='FF000000'),
                     top=Side(border_style='medium', color='FF000000'),
                     bottom=Side(border_style='medium', color='FF000000'))

        cell_border_sumas = Border(left=Side(border_style=None, color='FF000000'),
                     right=Side(border_style=None, color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))

        for cell in ws["1:1"]:
            cell.fill = PatternFill(start_color="8ccbff", end_color="8ccbff", fill_type = "solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

        for column_cells in ws.columns:
            #length = max(len(self.as_text(cell.value)) for cell in column_cells)
            length = len(self.as_text(column_cells[0].value))
            ws.column_dimensions[column_cells[0].column_letter].width = length+5

        ws.column_dimensions['A'].width = 12

        for cell in ws['A']:
            cell.font = Font(bold=True)


        for cell in ws[str(sumas_row)+":"+str(sumas_row)]:
            cell.border = cell_border_sumas
            cell.font = Font(bold=True)


        for i in range(2,sumas_row+1):
            for j in range(2,columna_totales+1):
                ws.cell(i,j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        ws.cell(sumas_row,1).fill = PatternFill(start_color="4ac6ff", end_color="4ac6ff", fill_type = "solid")
        ws.cell(sumas_row,1).border = cell_border


    def calculaAgregados(self, df, ws_cats, variable):
        meses = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
        por_categorias = df.groupby(['mes', 'tipo'], as_index=False).agg({variable:sum})
        por_categorias_wide = por_categorias.pivot_table(index="mes",columns=['tipo'],values=variable,fill_value= 0)
        por_categorias_wide.reset_index(inplace=True)
        por_categorias_wide['mes'] = pd.Categorical(por_categorias_wide['mes'], meses)
        por_categorias_wide = por_categorias_wide.sort_values("mes")
        por_categorias_wide['mes'] = por_categorias_wide.mes.astype(str)
        for r in dataframe_to_rows(por_categorias_wide, index=False, header=True):
            ws_cats.append(r)

        print(por_categorias_wide)
        self.numeroDeColumnas = len(por_categorias_wide.columns)
        self.columna_totales = self.numeroDeColumnas + 1
        self.sumas_row = len(por_categorias_wide.index)+2
        ws_cats.cell(1,self.columna_totales, "Total")
        ws_cats.cell(self.sumas_row,1,"Anual")

        for i in range(2,self.columna_totales):
            letra = get_column_letter(i)
            ws_cats.cell(self.sumas_row,i,"=SUM("+letra+ "2:"+letra+ str(self.sumas_row-1)+")")

        letra_final = get_column_letter(self.numeroDeColumnas)
        for i in range(2,self.sumas_row):
            ws_cats.cell(i,self.columna_totales,"=SUM(B"+str(i)+ ":"+letra_final+ str(i)+")")

        letra_sumas = get_column_letter(self.columna_totales)
        ws_cats.cell(self.sumas_row,self.columna_totales,"=SUM("+letra_sumas+"2:"+letra_sumas+str(self.sumas_row-1)  +")")

    def hazAgregados(self,paths):
        print(paths[0])
        year_folder = os.path.split(paths[0])[0]
        print(year_folder)
        year = os.path.split(os.path.split(paths[0])[0])[1]
        client = os.path.split(os.path.split(os.path.split(paths[0])[0])[0])[1]
        print(client+"_"+year)

        xlsx_path = os.path.join(year_folder, client+"_"+year + ".xlsx")

        workbook = Workbook()
        ws_todos = workbook.create_sheet("todos")
        sheet1_name = workbook.get_sheet_names()[0]
        sheet1 = workbook[sheet1_name]
        workbook.remove_sheet(sheet1)
        ws_cats = workbook.create_sheet("iva_por_categria")
        ws_cats_importe = workbook.create_sheet("importe_por_categria")



        ws_todos.cell(1, 1, "mes")
        ws_todos.cell(1, 2, 'clave_concepto')
        ws_todos.cell(1, 3, 'UUID')
        ws_todos.cell(1, 4, 'cantidad')
        ws_todos.cell(1, 5, 'importeConcepto')
        ws_todos.cell(1, 6, 'descripcion')
        ws_todos.cell(1, 7, 'impuestos')
        ws_todos.cell(1, 8, 'tipo')

        row = 1
        # for mes in meses:
        #     if mes in meses_folders:
        #         for concepto in self.conceptos[mes]:
        for concepto in self.conceptos:
            row += 1
            ws_todos.cell(row, 1, concepto['mes'])
            ws_todos.cell(row, 2, concepto['clave_concepto'])
            ws_todos.cell(row, 3, concepto['UUID'])
            ws_todos.cell(row, 4, concepto['cantidad'])
            ws_todos.cell(row, 5, concepto['importeConcepto'])
            ws_todos.cell(row, 6, concepto['descripcion'])
            ws_todos.cell(row, 7, concepto['impuestos'])
            ws_todos.cell(row, 8, concepto['tipo'])

        df = pd.DataFrame(self.conceptos)



        self.calculaAgregados(df, ws_cats, 'impuestos')
        self.style_ws(ws_cats, self.columna_totales, self.sumas_row)

        self.calculaAgregados(df, ws_cats_importe, 'importeConcepto')
        self.style_ws(ws_cats_importe, self.columna_totales, self.sumas_row)


        workbook.save(xlsx_path)



    def hazResumenDiot(self,currentDir):

        xlsx_path = os.path.join(currentDir,os.path.join("huiini","resumen.xlsx"))
        #workbook = xlsxwriter.Workbook(xlsx_path)
        #worksheet = workbook.add_worksheet("por_RFC")
        workbook = Workbook()
        ws_rfc = workbook.create_sheet("por_RFC")
        sheet1_name = workbook.get_sheet_names()[0]
        sheet1 = workbook[sheet1_name]
        workbook.remove_sheet(sheet1)

        ws_rfc.cell(1, 1,     "RFC")
        ws_rfc.cell(1, 2,     "SUBTOTAL")
        ws_rfc.cell(1, 3,     "DESCUENTO")
        ws_rfc.cell(1, 4,     "IMPORTE")
        ws_rfc.cell(1, 5,     "IVA")
        ws_rfc.cell(1, 6,     "TOTAL")

        row = 1
        for key, value in self.diccionarioPorRFCs.items():
            row += 1
            ws_rfc.cell(row, 1, key)
            ws_rfc.cell(row, 2, value['subTotal'])
            ws_rfc.cell(row, 3, value['descuento'])
            ws_rfc.cell(row, 4, value['trasladoIVA'])
            ws_rfc.cell(row, 5, value['importe'])
            ws_rfc.cell(row, 6, value['total'])

        #worksheet2 = workbook.add_worksheet("por_Factura")
        ws_factura = workbook.create_sheet("por_Factura")
        ws_factura.cell(1, 1, "clave_ps")
        ws_factura.cell(1, 2,     "Fecha")
        ws_factura.cell(1, 3,     "UUID")
        ws_factura.cell(1, 4,     "Nombre")
        ws_factura.cell(1, 5,     "RFC")
        ws_factura.cell(1, 6,     "Concepto")
        ws_factura.cell(1, 7,     "Sub")
        ws_factura.cell(1, 8,     "IVA")
        ws_factura.cell(1, 9,     "Total")
        ws_factura.cell(1, 10,     "F-Pago")
        ws_factura.cell(1, 11,     "M-Pago")
        ws_factura.cell(1, 12,     "Tipo")

        row = 1
        for factura in self.listaDeFacturasOrdenadas:
            row += 1
            ws_factura.cell(row, 1, factura.conceptos[0]['clave_concepto'])
            ws_factura.cell(row, 2, factura.fechaTimbrado)
            ws_factura.cell(row, 3, factura.UUID)
            ws_factura.cell(row, 4, factura.EmisorNombre)
            ws_factura.cell(row, 5, factura.EmisorRFC)
            ws_factura.cell(row, 6, factura.conceptos[0]['descripcion'])
            ws_factura.cell(row, 7, factura.subTotal)
            ws_factura.cell(row, 8, factura.traslados["IVA"]["importe"])
            ws_factura.cell(row, 9, factura.total)
            ws_factura.cell(row, 10, factura.formaDePagoStr)
            ws_factura.cell(row, 11, factura.metodoDePago)
            ws_factura.cell(row, 12, factura.conceptos[0]['tipo'])

        row += 1
        ws_factura.cell(row, 7,     "=SUM(G2:G"+str(row-1)+")")
        ws_factura.cell(row, 8,     "=SUM(H2:H"+str(row-1)+")")
        ws_factura.cell(row, 9,     "=SUM(I2:I"+str(row-1)+")")

        workbook.save(xlsx_path)








        #url_get = "http://huiini.pythonanywhere.com/resumen"


        # r = requests.get(url_get, stream=True,
        #                 auth=(self.w.username.text(), self.w.password.text()))
        # time_old.sleep(1)
        # if r.status_code == 200:
        #     with open(join(join(self.esteFolder,"huiini"), 'resumenDiot.xlsx'),'wb') as f:
        #         r.raw.decode_content = True
        #         shutil.copyfileobj(r.raw, f)



    def hazListadeUuids(self):
        self.listadeUuids = []
        for renglon in range(self.numeroDeFacturasValidas):
            self.listadeUuids.append(self.tableWidget_xml.item(renglon,1).text())


    def handleHeaderMenu(self, pos):
        menu = QtGui.QMenu()
        deleteAction = QtGui.QAction('&Delete', self)
        #deleteAction = QtGui.QAction("Delete")
        deleteAction.triggered.connect(lambda: self.quitaRenglon(self.tableWidget_xml.verticalHeader().logicalIndexAt(pos)))
        menu.addAction(deleteAction)

        menu.exec_(QtGui.QCursor.pos())

    def quitaRenglon(self,row):
        elNombre = self.tableWidget_xml.item(row,2).text()
        suRFC = ""
        for factura in self.listaDeFacturasOrdenadas:
            if factura.UUID == elNombre:
                print("i found it!")
                suRFC = factura.EmisorRFC

                break


        suSubtotal = float(self.tableWidget_xml.item(row,7).text())
        suDescuento = float(self.tableWidget_xml.item(row,8).text())
        suTrasladoIVA = float(self.tableWidget_xml.item(row,9).text())
        suImporte = float(self.tableWidget_xml.item(row,7).text())-float(self.tableWidget_xml.item(row,8).text())
        self.tableWidget_xml.removeRow(row)

        if suRFC in self.diccionarioPorRFCs:
            self.diccionarioPorRFCs[suRFC]['subTotal'] -= suSubtotal
            self.diccionarioPorRFCs[suRFC]['descuento'] -= suDescuento
            self.diccionarioPorRFCs[suRFC]['trasladoIVA'] -= suTrasladoIVA
            self.diccionarioPorRFCs[suRFC]['importe'] -= suImporte

            if math.fabs(self.diccionarioPorRFCs[suRFC]['subTotal']) < 0.0001 and math.fabs(self.diccionarioPorRFCs[suRFC]['descuento']) < 0.0001 and math.fabs(self.diccionarioPorRFCs[suRFC]['trasladoIVA']) < 0.0001 and math.fabs(self.diccionarioPorRFCs[suRFC]['importe']) < 0.0001:
                self.diccionarioPorRFCs.pop(suRFC,0)


        self.numeroDeFacturasValidas -= 1
        self.sumale(1)

        url_get =  "%s/remove/%s/%s" % (url_server, self.hash_carpeta, elNombre)

        r = requests.get(url_get, stream=True,
                        auth=(self.w.username.text(), self.w.password.text()))


        self.hazResumenDiot(self.esteFolder)
        # try:
        #     if os.path.exists(os.path.join(os.path.join(self.esteFolder,"huiini"),"resumenDiot.pdf")):
        #
        #         os.remove(os.path.join(os.path.join(self.esteFolder,"huiini"),"resumenDiot.pdf"))
        #
        #     os.rename(os.path.join(self.esteFolder,"resumenDiot.pdf"), os.path.join(os.path.join(self.esteFolder,"huiini"),"resumenDiot.pdf"))
        # except:
        #     QtGui.QMessageBox.information(self, "Information", "tienes abierto el resumenDiot.pdf")


    def sumale(self, renglonResumen=0):
        for columna in range(7,14):
            suma = 0
            for renglon in range(self.numeroDeFacturasValidas):

                suma += float(self.tableWidget_xml.item(renglon, columna).text())


            self.tableWidget_resumen.setItem(renglonResumen,columna-4,QTableWidgetItem(str(suma)))

        if renglonResumen == 1:
            self.tableWidget_resumen.setItem(0,1,QTableWidgetItem("            ---------"))
            self.tableWidget_resumen.setItem(0,2,QTableWidgetItem("Sumatoria del Periodo Original"))
            self.tableWidget_resumen.setItem(1,1,QTableWidgetItem("Resumen Diot Actualizado"))
            self.tableWidget_resumen.setItem(1,2,QTableWidgetItem("Sumatoria del Periodo Actualizada"))
            self.tableWidget_resumen.setCellWidget(1,0,ImgWidgetPalomita(self))
            self.tableWidget_resumen.setCellWidget(0,0,ImgWidgetTache(self))


    def ponEncabezado(self):
        itemVersion = QTableWidgetItem("V")
        itemVersion.setToolTip("Versión")
        self.tableWidget_xml.setHorizontalHeaderItem (0, QTableWidgetItem("Pdf"))
        self.tableWidget_xml.setHorizontalHeaderItem (1, QTableWidgetItem("Fecha"))
        self.tableWidget_xml.setHorizontalHeaderItem (2, QTableWidgetItem("UUID"))
        self.tableWidget_xml.setHorizontalHeaderItem (3, QTableWidgetItem("Receptor"))
        self.tableWidget_xml.setHorizontalHeaderItem (4, QTableWidgetItem("Emisor"))
        self.tableWidget_xml.setHorizontalHeaderItem (5, QTableWidgetItem("Concepto"))
        self.tableWidget_xml.setHorizontalHeaderItem (6, itemVersion)
        self.tableWidget_xml.setHorizontalHeaderItem (7, QTableWidgetItem("Subtotal"))
        self.tableWidget_xml.setHorizontalHeaderItem (8, QTableWidgetItem("Descuento"))
        self.tableWidget_xml.setHorizontalHeaderItem (9, QTableWidgetItem("Traslado\nIVA"))
        self.tableWidget_xml.setHorizontalHeaderItem (10, QTableWidgetItem("Traslado\nIEPS"))
        self.tableWidget_xml.setHorizontalHeaderItem (11, QTableWidgetItem("Retención\nIVA"))
        self.tableWidget_xml.setHorizontalHeaderItem (12, QTableWidgetItem("Retención\nISR"))
        self.tableWidget_xml.setHorizontalHeaderItem (13, QTableWidgetItem("Total"))
        self.tableWidget_xml.setHorizontalHeaderItem (14, QTableWidgetItem("Forma\nPago"))
        self.tableWidget_xml.setHorizontalHeaderItem (15, QTableWidgetItem("Método\nPago"))



    def meDoblePicaronXML(self, row,column):
        print("me picaron en : " +str(row)+", " +str(column))
#         if column == 5:
#             suUUID = self.tableWidget_xml.item(row,2).text()
#             laFactura = None
#             for factura in self.listaDeFacturasOrdenadas:
#                 if factura.UUID == suUUID:
#                     print("i found it!")
#                     laFactura = factura
#
#                     break
#             mesage = ""
#             for concepto in laFactura.conceptos:
#                 mesage += concepto["descripcion"] + u'\n'
#
#             QtGui.QMessageBox.information(self, "Conceptos", mesage)
        if column == 2:


            xml =join(self.esteFolder + os.sep,self.tableWidget_xml.item(row, 2).text()+".xml")
            #acrobatPath = r'C:/Program Files (x86)/Adobe/Acrobat Reader DC/Reader/AcroRd32.exe'
            #subprocess.Popen("%s %s" % (acrobatPath, pdf))
            try:
                os.startfile(xml)
                print("este guey me pico:"+xml)
            except:
                print ("el sistema no tiene una aplicacion por default para abrir xmls")
                QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir xmls" )

        if column == 0:

            pdf = join(join(self.esteFolder,"huiini"),self.tableWidget_xml.item(row, 2).text()+".pdf")
            #acrobatPath = r'C:/Program Files (x86)/Adobe/Acrobat Reader DC/Reader/AcroRd32.exe'
            #subprocess.Popen("%s %s" % (acrobatPath, pdf))
            try:
                os.startfile(pdf)
                print("este guey me pico:"+pdf)
            except:
                print ("el sistema no tiene una aplicacion por default para abrir pdfs")
                QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir pdfs" )


    def meDoblePicaronResumen(self, row,column):
        print("me picaron en : " +str(row)+", " +str(column))
        excel = join(join(self.esteFolder,"huiini"),"resumen.xlsx")
        #acrobatPath = r'C:/Program Files (x86)/Adobe/Acrobat Reader DC/Reader/AcroRd32.exe'
        #subprocess.Popen("%s %s" % (acrobatPath, pdf))
        try:
            os.startfile(excel)
            print("este guey me pico:"+excel)
        except:
            print ("el sistema no tiene una aplicacion por default para abrir exceles")
            QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir exceles" )

    def cambiaSeleccionDeImpresora(self, curr, prev):
        print(curr.text())
        self.impresoraDefault = curr.text()
        win32print.SetDefaultPrinter(self.impresoraDefault)

    def cambiaImpresora(self):
        # self.tabWidget.setCurrentIndex(1)
        self.listaDeImpresoras.setEnabled(True)

        for (a,b,name,d) in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL):
            self.listaDeImpresoras.addItem(name)

    def cancelaImpresion(self):
        print("cancelaria")



    def imprime(self):
        #objetosMagicosOrdenados = sorted(self.objetosMagicos, key=lambda objetosMagicos: objetosMagicos.fecha)

        for factura in self.listaDeFacturasOrdenadas:
            try:
                if factura.total > 0:
                    print(factura.fechaTimbrado)
                    hh = win32api.ShellExecute(0, "print", join(join(self.esteFolder,"huiini"), factura.UUID+".pdf"),None, ".",  0)
                    if hh > 40:
                        print("algo")
                        time_old.sleep(10)

                elif factura.total < 0:
                    print("negativo?????")
                else:#si es cero
                    print("nada")
            except:
                print("hay un pdf faltante o corrupto")


        hh = win32api.ShellExecute(0, "print", join(join(self.esteFolder,"huiini"), "resumenDiot.pdf") , None,  ".",  0)
    def esteItem(self, text, tooltip):
        item = QTableWidgetItem(text)
        item.setToolTip(tooltip)
        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
        return item


    def hazPDFs(self):
        contador = -1
        for factura in self.listaDeFacturasOrdenadas:
            contador += 1
            if factura.has_pdf == False:
                xml_name = basename(factura.xml_path)
                factura.conviertemeEnTex()
                factura.conviertemeEnPDF()


                factura.has_pdf = True
                self.tableWidget_xml.setCellWidget(contador,0, ImgWidgetPalomita(self))

                #     else:
                #         self.tableWidget_xml.setCellWidget(contador,0, ImgWidgetTache(self))
                # except:
                #     self.tableWidget_xml.setCellWidget(contador,0, ImgWidgetTache(self))
    def borraAuxiliares(self):
        self.pd.show()
        self.pd.setLabelText("Borrando Auxiliares...")
        contador = 0
        for t in range(0,10):
            time_old.sleep((1.0*len(self.listaDeFacturasOrdenadas)/10.0))
            self.pd.setValue((100.0 * contador) /(len(self.listaDeFacturasOrdenadas)*5.0))
        contador = 0
        for archivo in os.listdir(self.esteFolder):
            if ".tex" in archivo:
                contador += 1
                self.pd.setValue((100.0 * contador) /(len(self.listaDeFacturasOrdenadas)*5.0))
                eltex = join(self.esteFolder + os.sep,archivo)
                os.remove(eltex)
        for archivo in os.listdir(join(self.esteFolder,"huiini")):
            if ".log" in archivo:
                contador += 1
                self.pd.setValue((100.0 * contador) /(len(self.listaDeFacturasOrdenadas)*5.0))
                ellog = join(join(self.esteFolder,"huiini"),archivo)
                os.remove(ellog)
        for archivo in os.listdir(join(self.esteFolder,"huiini")):
            if ".aux" in archivo:
                contador += 1
                self.pd.setValue((100.0 * contador) /(len(self.listaDeFacturasOrdenadas)*5.0))
                elaux = join(join(self.esteFolder,"huiini"),archivo)
                os.remove(elaux)
        self.pd.hide()

    def cualCarpeta(self):
        self.folder.hide()
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.DirectoryOnly)
        file_dialog.setOption(QFileDialog.DontUseNativeDialog, True)
        file_view = file_dialog.findChild(QListView, 'listView')

        # to make it possible to select multiple directories:
        if file_view:
            file_view.setSelectionMode(QAbstractItemView.MultiSelection)
        f_tree_view = file_dialog.findChild(QTreeView)
        if f_tree_view:
            f_tree_view.setSelectionMode(QAbstractItemView.MultiSelection)

        if file_dialog.exec():
            paths = file_dialog.selectedFiles()

        self.procesaCarpetas(paths)

    def procesaCarpetas(self,paths):
        self.conceptos = []
        for path in paths:
            self.esteFolder = join(path,"EGRESOS")

            self.mes = os.path.split(path)[1]
            self.mes = ''.join([i for i in self.mes if not i.isdigit()])
            self.mes = self.mes.strip()
            #self.conceptos[self.mes] = []
            if not os.path.exists(join(self.esteFolder, "huiini")):
                os.makedirs(join(self.esteFolder, "huiini"))
            self.tableWidget_xml.clear()
            self.tableWidget_resumen.clear()
            self.tableWidget_resumen.repaint()
            self.ponEncabezado()
            self.tableWidget_xml.setRowCount(13)
            self.tableWidget_xml.repaint()
            cuantosDuplicados = 0
            self.listaDeDuplicados=[]
            self.listaDeFacturas = []
            self.listaDeUUIDs = []
            contador = 0
            for archivo in os.listdir(self.esteFolder):
                if ".xml" in archivo:

                    laFactura = Factura(join(self.esteFolder + os.sep,archivo))
                    if laFactura.sello == "SinSello":
                        print("Omitiendo xml sin sello "+laFactura.xml_path)
                    else:
                        if laFactura.version:
                            if laFactura.UUID in self.listaDeUUIDs:

                                cuantosDuplicados+=1
                                self.listaDeDuplicados.append(laFactura.UUID)
                            else:
                                self.listaDeUUIDs.append(laFactura.UUID)
                                contador += 1
                                self.listaDeFacturas.append(laFactura)

            if contador > 13:
                self.tableWidget_xml.setRowCount(contador)

            self.listaDeFacturasOrdenadas = sorted(self.listaDeFacturas, key=lambda listaDeFacturas: listaDeFacturas.fechaTimbrado)
            self.diccionarioPorRFCs = {}
            print(self.listaDeFacturasOrdenadas)


            self.pd =  QProgressDialog("Operation in progress.", "Cancel", 0, 100, self)
            self.pd.setWindowTitle("Huiini")
            self.pd.setValue(0)
            self.pd.show()

            if cuantosDuplicados > 0:
                mensaje = "hay "+str(cuantosDuplicados)+" duplicados\n"
                chunks = []
                for esteDuplicado in self.listaDeDuplicados:
                    chunks.append(str(esteDuplicado)+"\n")
                mensaje2 = "".join(chunks)
                mensaje = mensaje + mensaje2
                QMessageBox.information(self, "Information", mensaje)

            contador = 0
            for factura in self.listaDeFacturasOrdenadas:
                self.pd.setValue(50*((contador + 1)/len(self.listaDeFacturasOrdenadas)))
                factura.setFolio(contador + 1)
                los_conceptos = factura.conceptos.copy()
                for concepto in los_conceptos:
                    concepto["mes"] = self.mes
                    concepto["UUID"] = factura.UUID
                    if concepto["impuestos"]:
                        concepto["impuestos"] = float(concepto['impuestos'])
                    else:
                        concepto["impuestos"] = 0
                self.conceptos.extend(los_conceptos)
                self.pd.setLabelText("Procesando: " + factura.UUID[:17] + "...")

                #url = "http://huiini.pythonanywhere.com/upload"
                #url =  "%s/upload/%s/" % (url_server, self.hash_carpeta)

                ####################################################Definir puerto  80 80   ################################
                xml_path = factura.xml_path

                #xml_path = 'C:/Users/SICAD/Dropbox/Araceli/2017/JUNIO/EGRESOS/DE820CD4-2F37-4751-9D38-0FD6947CB287.xml'
                files = {'files': open(xml_path , 'rb')}
                # print(r.content
                # print(r.text)


                self.tableWidget_xml.setItem(contador,1,self.esteItem(factura.fechaTimbrado,factura.fechaTimbrado))
                self.tableWidget_xml.setItem(contador,2,self.esteItem(factura.UUID,factura.UUID))
                self.tableWidget_xml.setItem(contador,3,self.esteItem(factura.ReceptorRFC,factura.ReceptorNombre))
                self.tableWidget_xml.setItem(contador,4,self.esteItem(factura.EmisorRFC,factura.EmisorNombre))
                mesage = ""
                for concepto in factura.conceptos:
                    mesage += concepto["descripcion"] + u'\n'
                self.tableWidget_xml.setItem(contador,5, self.esteItem(factura.conceptos[0]['descripcion'],mesage))
                self.tableWidget_xml.setItem(contador,6,self.esteItem(str(factura.version),""))
                self.tableWidget_xml.setItem(contador,7,self.esteItem(str(factura.subTotal),""))
                self.tableWidget_xml.setItem(contador,8,self.esteItem(str(factura.descuento),""))
                self.tableWidget_xml.setItem(contador,9,self.esteItem(str(factura.traslados["IVA"]["importe"]),""))
                self.tableWidget_xml.setItem(contador,10,self.esteItem(str(factura.traslados["IEPS"]["importe"]),""))
                self.tableWidget_xml.setItem(contador,11,self.esteItem(str(factura.retenciones["IVA"]),""))
                self.tableWidget_xml.setItem(contador,12,self.esteItem(str(factura.retenciones["ISR"]),""))
                self.tableWidget_xml.setItem(contador,13,self.esteItem(str(factura.total),""))
                self.tableWidget_xml.setItem(contador,14,self.esteItem(factura.formaDePagoStr,""))
                self.tableWidget_xml.setItem(contador,15, self.esteItem(factura.metodoDePago,factura.metodoDePago))

                if factura.EmisorRFC in self.diccionarioPorRFCs:
                    self.diccionarioPorRFCs[factura.EmisorRFC]['subTotal'] += float(factura.subTotal)
                    self.diccionarioPorRFCs[factura.EmisorRFC]['descuento'] += float(factura.descuento)
                    self.diccionarioPorRFCs[factura.EmisorRFC]['trasladoIVA'] += float(factura.traslados['IVA']['importe'])
                    self.diccionarioPorRFCs[factura.EmisorRFC]['importe'] += float(factura.subTotal)-float(factura.descuento)
                    self.diccionarioPorRFCs[factura.EmisorRFC]['total'] += float(factura.total)
                    print("sumale " + str(factura.subTotal) )
                else:
                    self.diccionarioPorRFCs[factura.EmisorRFC] = {'subTotal': float(factura.subTotal),
                                                                  'descuento': float(factura.descuento),
                                                                  'trasladoIVA': float(factura.traslados['IVA']['importe']),
                                                                  'importe': float(factura.subTotal)-float(factura.descuento),
                                                                  'total': float(factura.total)
                                                                }
                    print("crealo con " + str(factura.subTotal))

                contador +=1

                # try:
                #     r = requests.post (url, files=files,
                #                        timeout=20,
                #                        data={'folio' :contador + 1},
                #                        auth=(self.w.username.text(), self.w.password.text()))
                # except:
                #     continue


            #if contador == len(self.listaDeFacturasOrdenadas):

            self.pd.show()
            self.pd.setLabelText("Creando Resumen...")
            for t in range(0,5):
                time_old.sleep(0.05*len(self.listaDeFacturasOrdenadas))
                self.pd.setValue(self.pd.value() + ( (100 - self.pd.value()) / 2))




            #self.hazPDFs()



            contador = -1

            # time_old.sleep(0.5*len(self.listaDeFacturasOrdenadas))

            self.imprimir.setEnabled(True)

            self.numeroDeFacturasValidas = len(self.listaDeFacturasOrdenadas)


            self.sumale()
            self.pd.setLabelText("Carpeta procesada")
            self.pd.setValue(self.pd.value() + ( (100 - self.pd.value()) / 2))
            self.hazResumenDiot(self.esteFolder)
            self.pd.setValue(100)
            self.tableWidget_resumen.setItem(0,1,QTableWidgetItem("Resumen Diot"))
            self.tableWidget_resumen.setItem(0,2,QTableWidgetItem("Sumatoria del Periodo"))
            self.tableWidget_resumen.setCellWidget(0,0, ImgWidgetPalomita(self))

            #obtener los warnings de las facturas
            mensajeAlerta =""
            for factura in self.listaDeFacturasOrdenadas:
                if not factura.mensaje == "":
                    mensajeAlerta += factura.UUID + factura.mensaje + r'\n'
            if not mensajeAlerta == "":
                QMessageBox.information(self, "Information", mensajeAlerta)




            self.folder.setText("Carpeta Procesada: " + u'\n' + self.esteFolder)
            self.folder.show()

        #self.borraAuxiliares()
        self.hazAgregados(paths)

        self.raise_()
        self.activateWindow()


app = QtWidgets.QApplication(sys.argv)
form = Ui_MainWindow()
form.show()


app.exec_()
